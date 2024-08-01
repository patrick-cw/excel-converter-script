import pandas as pd
from openpyxl import load_workbook
import sys
import os
import re
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import time

class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.logfile = open(filename, "a")

    def write(self, message):
        self.terminal.write(message)
        self.logfile.write(message)

    def flush(self):
        pass

# Redirect stdout to both console and file
sys.stdout = Logger("console.log")

def show_notification(title, message):
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Create a big notification pop-up
    root.lift()
    root.attributes("-topmost", True)
    messagebox.showinfo(title, message)
    root.destroy()

# Extract data from CSV
def extract_data(csv_file):
    #Get Team ID
    team = pd.read_csv(csv_file, nrows=1, usecols=[0], header=None).iloc[0, 0]
    team_all = team.split(' - ')
    pattern = r'CMT\d+'
    team_id = re.findall(pattern, team)
    
    if len(team_id)<1:
        raise Exception("Team ID is not found")
    #Read Data from CSV file
    df = pd.read_csv(csv_file,skiprows=1, skipfooter=1,engine='python')

    #Format some of the data
    columns_to_check = ['Present', 'Absent','Late','Total Present','Total Absent']
    missing_columns = [col for col in columns_to_check if col not in df.columns]
    if len(missing_columns) == 0:
        df.drop(['Present', 'Absent','Late','Total Present','Total Absent'], axis=1, inplace=True)
    else:
         df.drop(['Present', 'Absent'], axis=1, inplace=True)

    # Load the Excel file to check team and season based on team id
    df['Flex Team ID'] = team_id[0]
    team_file = "SourceFilename.xlsx"
    cm_sheet = "Team History" 
    all_data = pd.read_excel(team_file,sheet_name=cm_sheet)
    team_info = all_data.loc[all_data['Team ID']==team_id[0]]
    df['Season'] = team_all[1]
    df['Team'] = team_info["Team"].iloc[0]
    
    return df

def transform_data(data):
    # Transform data
    data = data.melt(id_vars=['Season','Team','Flex Team ID','Name'], var_name='Date', value_name='Attendance')
    data.rename(columns={'Name':'Player'}, inplace=True)
    #data['Date'] = pd.to_datetime(data['Date'], format='mixed', dayfirst=True)
    print(data['Date'].iloc[20:21])
    data['Date'] = pd.to_datetime(data['Date'], format='%a %m/%d/%y')
    data['Date'] = data['Date'].dt.date
    data.loc[data['Attendance']=='Late','Attendance'] = 'Present'

    return data

# Load data into Excel
def input_data_to_excel(data, excel_file):
    
    wb = load_workbook(excel_file)
    ws = wb.active
    column_names = ['Season','Team','Flex Team ID','Player','Date','Attendance']
    
    #Check Column Name is already made or not
    if not ws.cell(row=1, column=1).value:
        # Add column names
        for idx, column_name in enumerate(column_names, start=1):
            ws.cell(row=1, column=idx, value=column_name)
    next_row = ws.max_row + 1

    #Input the data on the next active row
    for index, row in data.iterrows():
        for col_idx, (key,value) in enumerate(row.items(), start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            #Change date format
            if key == 'Date':
                cell.number_format = 'dd/mm/yyyy'
        next_row += 1
    wb.save(excel_file)
    wb.close()
    return 0


def main():
    print("Run a New Task")
    start_all = time.time()
    start_time = time.time()
    csv_file = new_filename 
    filename = os.path.basename(new_filename)
    excel_file = 'TargetFileName.xlsx' 
    dest_sheet = "NewFormat" 
    print('1. Read Existing Data...')
    existing_df = pd.read_excel(excel_file,sheet_name=dest_sheet)
    end_time = time.time()
    total_time = "{:.2f}".format(end_time - start_time)
    print('Runtime:',total_time,'S')

    try:
        # Extract
        start_time = time.time()
        print('2. Extract Data')
        data = extract_data(csv_file)
        end_time = time.time()
        total_time = "{:.2f}".format(end_time - start_time)
        print('Runtime:',total_time,'S')

        # Transform
        start_time = time.time()
        print('3. Transform Data')
        data = transform_data(data)
        end_time = time.time()
        total_time = "{:.2f}".format(end_time - start_time)
        print('Runtime:',total_time,'S')
        #Check Duplicates
        len_input_origin = len(data)
        
        start_time = time.time()
        print('4. Check Duplicated Data')
        if len(existing_df)<1:
            column_names = ['Season', 'Team', 'Magic Team ID', 'Player', 'Date', 'Attendance']
            # Create an empty DataFrame with the specified column names
            existing_df = pd.DataFrame(columns=column_names)
            len_input_filtered = len_input_origin
            len_duplicate = 0
        else:
            existing_df['Date'] = existing_df['Date'].dt.date

            merged = pd.merge(existing_df, data, on=['Season', 'Team', 'Magic Team ID', 'Player', 'Date', 'Attendance'], how='inner')
            data = data[~data.isin(merged)].dropna()

            len_input_filtered = len(data)
            len_duplicate = len_input_origin - len_input_filtered
        end_time = time.time()
        total_time = "{:.2f}".format(end_time - start_time)
        print('Runtime:',total_time,'S')

        # Load

        start_time = time.time()
        print('5. Input Data to Excel')
        input_data_to_excel(data, excel_file)
        end_time = time.time()
        total_time = "{:.2f}".format(end_time - start_time)
        print('Runtime:',total_time,'S')

        #Check Double Entry with Different Attendances
        start_time = time.time()
        print('6. Check Data with Different Attendances Log')
        check_df = pd.read_excel(excel_file,sheet_name=dest_sheet)
        print(check_df)
        duplicate_groups = check_df.groupby(['Season', 'Team', 'Magic Team ID', 'Player', 'Date']).filter(lambda x: x['Attendance'].nunique() > 1).sort_values(by=['Season', 'Team', 'Magic Team ID', 'Player', 'Date'])
        row_numbers = duplicate_groups.index.tolist()
        end_time = time.time()
        total_time = "{:.2f}".format(end_time - start_time)
        print('Runtime:',total_time,'S')

        if len(row_numbers)>0 and len_duplicate>0:
            print('Task Success (Attendance Error & Duplicate Found)',f"Filename : {filename} \nDestination File : {excel_file} \n\nRows Inputted : {len_input_filtered} Rows \nDuplicated Data : {len_duplicate} Rows \nRow With Attendance Error: {row_numbers}")
            show_notification("Task Success (Attendance Error & Duplicate Found) ", f"Filename : {filename} \nDestination File : {excel_file} \n\nRows Inputted : {len_input_filtered} Rows \nDuplicated Data : {len_duplicate} Rows \nRow With Attendance Error: {row_numbers}")
        elif len(row_numbers)==0 and len_duplicate>0:
            print('Task Success (Duplicate Found)')
            show_notification("Task Success (Duplicate Found)", f"Filename : {filename} \nDestination File : {excel_file} \n\nRows Inputted : {len_input_filtered} Rows \nDuplicated Data : {len_duplicate} Rows \nRow With Attendance Error: {0} Rows")
        elif len(row_numbers)>0 and len_duplicate==0:
            print('Task Success (Attendance Error)')
            show_notification("Task Success (Attendance Error) ", f"Filename : {filename} \nDestination File : {excel_file} \n\nRows Inputted : {len_input_filtered} Rows \nDuplicated Data : {len_duplicate} Rows \nRow With Attendance Error: {row_numbers}")
        else:
            print('Task Success')
            show_notification("Task Success", f"Filename : {filename} \nDestination File : {excel_file} \n\nRows Inputted : {len_input_filtered} Rows \nDuplicated Data : {len_duplicate} Rows \nRow With Attendance Error: {0} Rows")
            
    except Exception as e:
        print("Task Error",f"Error Message : {str(e)}")
        show_notification("Task Error",f"Filename : {filename} \nError Message : {str(e)}")
    end_all = time.time()
    total_time_all = "{:.2f}".format(end_all - start_all)
    print('Total Runtime:',total_time_all,'S')
    print("")

if __name__ == "__main__":
    new_filename = sys.argv[1]
    main()
    